from __future__ import annotations

import csv
import io
import json
import re
import uuid
from dataclasses import dataclass
from collections.abc import AsyncIterator
from typing import Any, Iterable

import httpx
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import Barcode, InventoryItem, InventoryItemBarcode
from app.schemas.barcode import BarcodeCreate
from app.schemas.inventory_import import (
    InventoryImportConfirmResponse,
    InventoryImportItemData,
    InventoryImportPreviewResponse,
    InventoryImportPreviewRow,
)
from app.schemas.inventory_item import InventoryItemCreate, _coerce_inventory_status
from app.services import barcode_service
from app.services.barcode_generation_service import compute_ean13_check_digit


_HEADER_ALIASES: dict[str, str] = {
    "id": "id",
    "title": "title",
    "name": "title",
    "item": "title",
    "item_name": "title",
    "equipment": "title",
    "equipment_name": "title",
    "inventory": "title",
    "inventory_name": "title",
    # RU / KZ common headers
    "наименование": "title",
    "название": "title",
    "имя": "title",
    "предмет": "title",
    "оборудование": "title",
    "инвентарь": "title",
    "description": "description",
    "descriptiom": "description",
    "desc": "description",
    "описание": "description",
    "примечание": "description",
    "category": "category",
    "категория": "category",
    "location": "location",
    "location_id": "location_id",
    "room": "location",
    "room_id": "location_id",
    "кабинет": "location",
    "кабинет_id": "location_id",
    "аудитория": "location",
    "помещение": "location",
    "комната": "location",
    "responsible_username": "responsible_username",
    "responsible_email": "responsible_username",
    "email": "responsible_username",
    "e_mail": "responsible_username",
    "почта": "responsible_username",
    "эл_почта": "responsible_username",
    "элпочта": "responsible_username",
    "ответственный_email": "responsible_username",
    "reponsible_username": "responsible_username",
    "reponsible": "responsible_username",
    "reponsible_username_email": "responsible_username",
    "responsible_first_name": "responsible_first_name",
    "responsible_last_name": "responsible_last_name",
    "имя_ответственного": "responsible_first_name",
    "фамилия_ответственного": "responsible_last_name",
    "status": "status",
    "статус": "status",
    "barcode_id": "barcode_id",
    "barcode_data_12": "barcode_data_12",
    "barcode_value": "barcode_data_12",
    "barcode": "barcode_data_12",
    "штрихкод": "barcode_data_12",
    "barcode_data": "barcode_data_12",
}


def _normalize_header(raw: object) -> str:
    if raw is None:
        return ""
    value = str(raw).strip().lower()
    value = value.replace("\ufeff", "")
    value = re.sub(r"[\s\-]+", "_", value)
    # Keep unicode letters/digits (Cyrillic headers, etc.).
    value = re.sub(r"[^\w]", "", value, flags=re.UNICODE)
    return value


def _cell_to_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s != "" else None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip() or None


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    s = value.strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _normalize_barcode(value: str | None) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    raw = "".join(raw.split())
    if not raw:
        return None
    if not raw.isdigit():
        return raw
    if len(raw) == 12:
        return raw + compute_ean13_check_digit(raw)
    return raw


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _looks_like_email(value: str | None) -> bool:
    if value is None:
        return False
    return bool(_EMAIL_RE.match(value.strip().lower()))


def _coerce_status(value: str | None) -> object:
    if value is None:
        return None
    s = value.strip()
    if not s:
        return None
    lowered = s.lower()
    ru_map = {
        "новое": "NEW",
        "в ремонте": "IN_REPAIR",
        "отремонтировано": "REPAIRED",
        "списано": "WRITTEN_OFF",
        "на складе": "IN_STOCK",
        "выдано": "ISSUED",
    }
    if lowered in ru_map:
        return ru_map[lowered]
    return _coerce_inventory_status(s)


@dataclass(frozen=True)
class _ParsedRow:
    row_number: int
    data: InventoryImportItemData


def parse_import_file(*, filename: str, content: bytes) -> list[_ParsedRow]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xltx") or name.endswith(".xltm"):
        return _parse_xlsx(content)
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="unsupported_file_type",
    )


def _parse_csv(content: bytes) -> list[_ParsedRow]:
    text = content.decode("utf-8-sig", errors="replace")
    header_line = ""
    for line in text.splitlines():
        if line.strip():
            header_line = line
            break

    delimiter_candidates = [",", ";", "\t"]
    delimiter_scores = {d: header_line.count(d) for d in delimiter_candidates}
    delimiter = max(delimiter_scores, key=lambda d: delimiter_scores[d]) if header_line else "\t"
    if delimiter_scores.get(delimiter, 0) == 0:
        # fallback: try sniffing, but default to tab (Excel often exports TSV)
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="," + ";" + "\t")
            delimiter = getattr(dialect, "delimiter", "\t") or "\t"
        except Exception:
            delimiter = "\t"

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    mapping: dict[str, str] = {}
    for h in headers:
        normalized = _normalize_header(h)
        mapped = _HEADER_ALIASES.get(normalized)
        if mapped:
            mapping[h] = mapped

    rows: list[_ParsedRow] = []
    for idx, raw_row in enumerate(reader, start=2):  # header is row 1
        data_dict: dict[str, Any] = {}
        for raw_key, raw_value in raw_row.items():
            if raw_key is None:
                continue
            key = mapping.get(raw_key)
            if not key:
                continue
            data_dict[key] = _cell_to_str(raw_value)
        rows.append(_build_parsed_row(idx, data_dict))
    return rows


def _parse_xlsx(content: bytes) -> list[_ParsedRow]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="xlsx_support_not_installed",
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [(_cell_to_str(h) or "") for h in header_row]
    mapping_by_index: dict[int, str] = {}
    for i, h in enumerate(headers):
        normalized = _normalize_header(h)
        mapped = _HEADER_ALIASES.get(normalized)
        if mapped:
            mapping_by_index[i] = mapped

    parsed: list[_ParsedRow] = []
    excel_row_number = 1
    for excel_row_number, row in enumerate(rows_iter, start=2):
        data_dict: dict[str, Any] = {}
        for i, cell in enumerate(row):
            key = mapping_by_index.get(i)
            if not key:
                continue
            data_dict[key] = _cell_to_str(cell)
        parsed.append(_build_parsed_row(excel_row_number, data_dict))
    return parsed


def _build_parsed_row(row_number: int, data_dict: dict[str, Any]) -> _ParsedRow:
    # Normalize known types
    item_id = _parse_int(data_dict.get("id"))
    barcode_id = _parse_int(data_dict.get("barcode_id"))
    location_id = _parse_int(data_dict.get("location_id"))

    location_raw = data_dict.get("location")
    if location_id is None and location_raw is not None:
        candidate_id = _parse_int(location_raw)
        if candidate_id is not None:
            location_id = candidate_id
            location_raw = None

    status_value = _coerce_status(data_dict.get("status"))

    data = InventoryImportItemData(
        id=item_id,
        title=data_dict.get("title"),
        description=data_dict.get("description"),
        category=data_dict.get("category"),
        location=location_raw,
        location_id=location_id,
        responsible_username=data_dict.get("responsible_username"),
        responsible_first_name=data_dict.get("responsible_first_name"),
        responsible_last_name=data_dict.get("responsible_last_name"),
        status=status_value,  # validator coerces
        barcode_id=barcode_id,
        barcode_data_12=data_dict.get("barcode_data_12"),
    )
    return _ParsedRow(row_number=row_number, data=data)


async def build_preview(
    *,
    token: str,
    rows: list[_ParsedRow],
    db: Session,
) -> InventoryImportPreviewResponse:
    item_ids = [row.data.id for row in rows if isinstance(row.data.id, int)]
    existing_item_ids: set[int] = set()
    if item_ids:
        existing_item_ids = set(
            db.execute(select(InventoryItem.id).where(InventoryItem.id.in_(item_ids)))
            .scalars()
            .all()
        )

    barcode_ids = [row.data.barcode_id for row in rows if isinstance(row.data.barcode_id, int)]
    barcode_ids_in_use: set[int] = set()
    if barcode_ids:
        barcode_ids_in_use = set(
            db.execute(
                select(InventoryItem.barcode_id).where(InventoryItem.barcode_id.in_(barcode_ids))
            )
            .scalars()
            .all()
        )

    normalized_barcode_values_by_row: dict[int, str] = {}
    barcode_values: list[str] = []
    for row in rows:
        normalized = _normalize_barcode(row.data.barcode_data_12)
        if normalized:
            normalized_barcode_values_by_row[row.row_number] = normalized
            barcode_values.append(normalized)

    existing_barcodes_by_value: dict[str, Barcode] = {}
    if barcode_values:
        for barcode in db.execute(select(Barcode).where(Barcode.value.in_(barcode_values))).scalars().all():
            if barcode.value:
                existing_barcodes_by_value[barcode.value] = barcode

    barcode_values_in_use: set[str] = set()
    if existing_barcodes_by_value:
        barcode_ids_existing = [b.id for b in existing_barcodes_by_value.values()]
        in_use_ids = set(
            db.execute(select(InventoryItem.barcode_id).where(InventoryItem.barcode_id.in_(barcode_ids_existing)))
            .scalars()
            .all()
        )
        for value, barcode in existing_barcodes_by_value.items():
            if barcode.id in in_use_ids:
                barcode_values_in_use.add(value)

    # external resolution (read-only): rooms by name, users by email
    room_name_to_id: dict[str, int] = {}
    user_email_to_id: dict[str, int] = {}
    unique_room_names = sorted(
        {str(row.data.location).strip() for row in rows if row.data.location}
    )
    unique_user_emails = sorted(
        {
            str(row.data.responsible_username).strip()
            for row in rows
            if row.data.responsible_username and _looks_like_email(str(row.data.responsible_username))
        }
    )

    async with httpx.AsyncClient(timeout=20) as client:
        if unique_room_names:
            try:
                resp = await client.get(
                    f"{settings.location_service_url}/rooms",
                    headers={"Authorization": f"Bearer {token}"},
                )
                if resp.status_code == status.HTTP_200_OK:
                    data = resp.json()
                    if isinstance(data, list):
                        for item in data:
                            if not isinstance(item, dict):
                                continue
                            rid = item.get("id")
                            name = item.get("name")
                            if isinstance(rid, int) and isinstance(name, str) and name.strip():
                                room_name_to_id[name.strip().lower()] = rid
            except Exception:
                # preview should still work without external resolution
                pass

        for email in unique_user_emails:
            normalized_email = email.strip().lower()
            if not normalized_email:
                continue
            try:
                resp = await client.get(
                    f"{settings.auth_service_url}/auth/users/search",
                    headers={"Authorization": f"Bearer {token}"},
                    params={"q": normalized_email, "limit": 20, "offset": 0},
                )
                if resp.status_code != status.HTTP_200_OK:
                    continue
                data = resp.json()
                if not isinstance(data, list):
                    continue
                for item in data:
                    if not isinstance(item, dict):
                        continue
                    uid = item.get("id")
                    uemail = item.get("email")
                    if isinstance(uid, int) and isinstance(uemail, str) and uemail.strip().lower() == normalized_email:
                        user_email_to_id[normalized_email] = uid
                        break
            except Exception:
                continue

    seen_barcode_values: set[str] = set()
    duplicate_barcode_values: set[str] = set()
    for value in barcode_values:
        if value in seen_barcode_values:
            duplicate_barcode_values.add(value)
        seen_barcode_values.add(value)

    preview_rows: list[InventoryImportPreviewRow] = []
    to_create = 0
    skipped = 0
    errors_count = 0

    for row in rows:
        errors: list[str] = []
        warnings: list[str] = []

        data = row.data.model_copy(deep=True)

        if data.id is not None and data.id in existing_item_ids:
            preview_rows.append(
                InventoryImportPreviewRow(
                    row_number=row.row_number,
                    action="skip_existing",
                    data=data,
                    errors=[],
                    warnings=[],
                )
            )
            skipped += 1
            continue
        else:
            action = "create"

        if not data.title:
            errors.append("title_required")

        normalized_value = normalized_barcode_values_by_row.get(row.row_number)
        if normalized_value:
            data.barcode_data_12 = normalized_value
            if normalized_value in duplicate_barcode_values:
                errors.append("duplicate_barcode_in_file")
            if normalized_value in barcode_values_in_use:
                errors.append("barcode_value_already_in_use")
            elif normalized_value in existing_barcodes_by_value:
                warnings.append("barcode_value_exists_will_reuse")
            else:
                warnings.append("barcode_value_missing_will_create")

        if data.barcode_id is not None:
            if data.barcode_id in barcode_ids_in_use:
                errors.append("barcode_id_already_in_use")
            else:
                barcode = db.get(Barcode, data.barcode_id)
                if not barcode:
                    # If barcode_id is from an external export, prefer barcode_data_12 to create/reuse.
                    if normalized_value:
                        data.barcode_id = None
                        warnings.append("barcode_id_not_found_ignored_using_barcode_data_12")
                    else:
                        errors.append("barcode_id_not_found")

        if data.location_id is None and data.location:
            resolved = room_name_to_id.get(data.location.strip().lower())
            if resolved is not None:
                data.location_id = resolved
                warnings.append("location_resolved_by_name")
            else:
                warnings.append("location_missing_will_create")

        if data.responsible_id is None and data.responsible_username:
            if _looks_like_email(data.responsible_username):
                resolved = user_email_to_id.get(data.responsible_username.strip().lower())
                if resolved is not None:
                    data.responsible_id = resolved
                    warnings.append("responsible_resolved_by_email")
                else:
                    warnings.append("responsible_missing_will_create")
            else:
                warnings.append("responsible_username_not_email_ignored")

        if errors:
            action = "error"
            errors_count += 1
        elif action == "create":
            to_create += 1

        preview_rows.append(
            InventoryImportPreviewRow(
                row_number=row.row_number,
                action=action,
                data=data,
                errors=errors,
                warnings=warnings,
            )
        )

    return InventoryImportPreviewResponse(
        total_rows=len(rows),
        to_create_count=to_create,
        skip_count=skipped,
        error_count=errors_count,
        rows=preview_rows,
    )


async def confirm_import(
    *,
    token: str,
    rows: list[_ParsedRow],
    db: Session,
    create_missing_locations: bool = True,
    create_missing_users: bool = True,
) -> InventoryImportConfirmResponse:
    preview = await build_preview(token=token, rows=rows, db=db)

    created_item_ids: list[int] = []
    skipped_count = 0
    error_rows: list[dict[str, Any]] = []

    # Cache for this import run
    room_name_to_id: dict[str, int] = {}
    user_email_to_id: dict[str, int] = {}

    async with httpx.AsyncClient(timeout=30) as client:
        # preload rooms for name -> id resolution and idempotent creation
        try:
            resp = await client.get(
                f"{settings.location_service_url}/rooms",
                headers={"Authorization": f"Bearer {token}"},
            )
            if resp.status_code == status.HTTP_200_OK and isinstance(resp.json(), list):
                for item in resp.json():
                    if not isinstance(item, dict):
                        continue
                    rid = item.get("id")
                    name = item.get("name")
                    if isinstance(rid, int) and isinstance(name, str) and name.strip():
                        room_name_to_id[name.strip().lower()] = rid
        except Exception:
            pass

        for row in preview.rows:
            if row.action == "skip_existing":
                skipped_count += 1
                continue
            if row.action != "create":
                error_rows.append({"row_number": row.row_number, "detail": "row_not_importable", "errors": row.errors})
                continue

            data = row.data
            try:
                # Resolve / create location
                location_id = data.location_id
                if location_id is None and data.location:
                    key = data.location.strip().lower()
                    location_id = room_name_to_id.get(key)
                    if location_id is None and create_missing_locations:
                        resp = await client.post(
                            f"{settings.location_service_url}/rooms",
                            headers={"Authorization": f"Bearer {token}"},
                            json={"name": data.location.strip(), "room_type": "UNKNOWN"},
                        )
                        if resp.status_code != status.HTTP_201_CREATED:
                            raise HTTPException(
                                status_code=status.HTTP_502_BAD_GATEWAY,
                                detail="location_service_error",
                            )
                        created = resp.json()
                        if not isinstance(created, dict) or not isinstance(created.get("id"), int):
                            raise HTTPException(
                                status_code=status.HTTP_502_BAD_GATEWAY,
                                detail="location_service_invalid_response",
                            )
                        location_id = int(created["id"])
                        room_name_to_id[key] = location_id

                # Resolve / create responsible user
                responsible_id = data.responsible_id
                email = (
                    data.responsible_username.strip().lower()
                    if data.responsible_username and _looks_like_email(data.responsible_username)
                    else None
                )
                if responsible_id is None and email:
                    responsible_id = user_email_to_id.get(email)
                    if responsible_id is None:
                        resp = await client.get(
                            f"{settings.auth_service_url}/auth/users/search",
                            headers={"Authorization": f"Bearer {token}"},
                            params={"q": email, "limit": 20, "offset": 0},
                        )
                        if resp.status_code == status.HTTP_200_OK and isinstance(resp.json(), list):
                            for item in resp.json():
                                if not isinstance(item, dict):
                                    continue
                                uid = item.get("id")
                                uemail = item.get("email")
                                if isinstance(uid, int) and isinstance(uemail, str) and uemail.strip().lower() == email:
                                    responsible_id = uid
                                    break

                    if responsible_id is None and create_missing_users:
                        password = uuid.uuid4().hex + "Aa1!"
                        full_name = " ".join(
                            [p for p in [data.responsible_first_name, data.responsible_last_name] if p]
                        ).strip() or None
                        resp = await client.post(
                            f"{settings.auth_service_url}/admin/users",
                            headers={"Authorization": f"Bearer {token}"},
                            json={
                                "email": email,
                                "password": password,
                                "full_name": full_name,
                                "first_name": data.responsible_first_name,
                                "last_name": data.responsible_last_name,
                                "role": None,
                                "role_ids": [],
                                "is_active": True,
                            },
                        )
                        if resp.status_code != status.HTTP_201_CREATED:
                            raise HTTPException(
                                status_code=status.HTTP_502_BAD_GATEWAY,
                                detail="auth_service_error",
                            )
                        created = resp.json()
                        if not isinstance(created, dict) or not isinstance(created.get("id"), int):
                            raise HTTPException(
                                status_code=status.HTTP_502_BAD_GATEWAY,
                                detail="auth_service_invalid_response",
                            )
                        responsible_id = int(created["id"])

                    if responsible_id is not None:
                        user_email_to_id[email] = responsible_id

                # Resolve / create barcode
                barcode_id = data.barcode_id
                normalized_value = _normalize_barcode(data.barcode_data_12)

                if barcode_id is not None:
                    existing_by_id = db.get(Barcode, barcode_id)
                    if existing_by_id is None:
                        if normalized_value:
                            barcode_id = None
                        else:
                            raise HTTPException(
                                status_code=status.HTTP_404_NOT_FOUND,
                                detail="barcode_id_not_found",
                            )

                if barcode_id is None and normalized_value:
                    existing = db.execute(select(Barcode).where(Barcode.value == normalized_value)).scalar_one_or_none()
                    if existing is not None:
                        barcode_id = existing.id
                    else:
                        created = barcode_service.create_barcode(
                            BarcodeCreate(value=normalized_value[:12] if len(normalized_value) == 13 else normalized_value, title=data.title),
                            db,
                            commit=False,
                        )
                        barcode_id = created.id

                if barcode_id is not None:
                    conflict = (
                        db.execute(select(InventoryItem.id).where(InventoryItem.barcode_id == barcode_id))
                        .scalar_one_or_none()
                    )
                    if conflict is not None:
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail="barcode_already_in_use",
                        )

                payload = InventoryItemCreate(
                    title=data.title or "",
                    description=data.description,
                    category=data.category,
                    location_id=location_id,
                    responsible_id=responsible_id,
                    status=data.status,
                    barcode_id=barcode_id,
                )

                item = InventoryItem(**payload.model_dump())
                if data.id is not None:
                    item.id = data.id
                db.add(item)
                db.flush()

                if barcode_id is not None:
                    item.barcode_id = barcode_id
                    db.add(InventoryItemBarcode(inventory_item_id=item.id, barcode_id=barcode_id))
                else:
                    generated = barcode_service.create_barcode(
                        BarcodeCreate(value=None, title=item.title),
                        db,
                        commit=False,
                    )
                    item.barcode_id = generated.id
                    db.add(InventoryItemBarcode(inventory_item_id=item.id, barcode_id=generated.id))

                db.commit()
                db.refresh(item)
                created_item_ids.append(item.id)
            except HTTPException as exc:
                db.rollback()
                error_rows.append(
                    {
                        "row_number": row.row_number,
                        "detail": exc.detail,
                        "status_code": exc.status_code,
                    }
                )
            except Exception:
                db.rollback()
                error_rows.append({"row_number": row.row_number, "detail": "unexpected_error"})

    return InventoryImportConfirmResponse(
        created_count=len(created_item_ids),
        skipped_count=skipped_count,
        error_count=len(error_rows),
        created_item_ids=created_item_ids,
        errors=error_rows,
    )


def _sse(event: str, data: dict[str, Any]) -> bytes:
    json_data = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    return f"event: {event}\ndata: {json_data}\n\n".encode("utf-8")


def confirm_import_stream(
    *,
    token: str,
    rows: list[_ParsedRow],
    db: Session,
    create_missing_locations: bool = True,
    create_missing_users: bool = True,
) -> AsyncIterator[bytes]:
    async def _gen():
        preview = await build_preview(token=token, rows=rows, db=db)
        yield _sse(
            "init",
            {
                "total_rows": preview.total_rows,
                "to_create_count": preview.to_create_count,
                "skip_count": preview.skip_count,
                "error_count": preview.error_count,
            },
        )

        created_item_ids: list[int] = []
        skipped_count = 0
        error_rows: list[dict[str, Any]] = []

        room_name_to_id: dict[str, int] = {}
        user_email_to_id: dict[str, int] = {}

        async with httpx.AsyncClient(timeout=30) as client:
            try:
                resp = await client.get(
                    f"{settings.location_service_url}/rooms",
                    headers={"Authorization": f"Bearer {token}"},
                )
                if resp.status_code == status.HTTP_200_OK and isinstance(resp.json(), list):
                    for item in resp.json():
                        if not isinstance(item, dict):
                            continue
                        rid = item.get("id")
                        name = item.get("name")
                        if isinstance(rid, int) and isinstance(name, str) and name.strip():
                            room_name_to_id[name.strip().lower()] = rid
            except Exception:
                pass

            total = len(preview.rows)
            for index, row in enumerate(preview.rows, start=1):
                if row.action == "skip_existing":
                    skipped_count += 1
                    yield _sse(
                        "row",
                        {
                            "index": index,
                            "total": total,
                            "row_number": row.row_number,
                            "result": "skipped",
                            "title": row.data.title,
                        },
                    )
                    continue

                if row.action != "create":
                    error_rows.append(
                        {
                            "row_number": row.row_number,
                            "detail": "row_not_importable",
                            "errors": row.errors,
                        }
                    )
                    yield _sse(
                        "row",
                        {
                            "index": index,
                            "total": total,
                            "row_number": row.row_number,
                            "result": "error",
                            "detail": "row_not_importable",
                            "errors": row.errors,
                            "title": row.data.title,
                        },
                    )
                    continue

                data = row.data
                try:
                    location_id = data.location_id
                    if location_id is None and data.location:
                        key = data.location.strip().lower()
                        location_id = room_name_to_id.get(key)
                        if location_id is None and create_missing_locations:
                            resp = await client.post(
                                f"{settings.location_service_url}/rooms",
                                headers={"Authorization": f"Bearer {token}"},
                                json={"name": data.location.strip(), "room_type": "UNKNOWN"},
                            )
                            if resp.status_code != status.HTTP_201_CREATED:
                                raise HTTPException(
                                    status_code=status.HTTP_502_BAD_GATEWAY,
                                    detail="location_service_error",
                                )
                            created = resp.json()
                            if not isinstance(created, dict) or not isinstance(created.get("id"), int):
                                raise HTTPException(
                                    status_code=status.HTTP_502_BAD_GATEWAY,
                                    detail="location_service_invalid_response",
                                )
                            location_id = int(created["id"])
                            room_name_to_id[key] = location_id

                    responsible_id = data.responsible_id
                    email = (
                        data.responsible_username.strip().lower()
                        if data.responsible_username and _looks_like_email(data.responsible_username)
                        else None
                    )
                    if responsible_id is None and email:
                        responsible_id = user_email_to_id.get(email)
                        if responsible_id is None:
                            resp = await client.get(
                                f"{settings.auth_service_url}/auth/users/search",
                                headers={"Authorization": f"Bearer {token}"},
                                params={"q": email, "limit": 20, "offset": 0},
                            )
                            if resp.status_code == status.HTTP_200_OK and isinstance(resp.json(), list):
                                for item in resp.json():
                                    if not isinstance(item, dict):
                                        continue
                                    uid = item.get("id")
                                    uemail = item.get("email")
                                    if (
                                        isinstance(uid, int)
                                        and isinstance(uemail, str)
                                        and uemail.strip().lower() == email
                                    ):
                                        responsible_id = uid
                                        break

                        if responsible_id is None and create_missing_users:
                            password = uuid.uuid4().hex + "Aa1!"
                            full_name = " ".join(
                                [
                                    p
                                    for p in [
                                        data.responsible_first_name,
                                        data.responsible_last_name,
                                    ]
                                    if p
                                ]
                            ).strip() or None
                            resp = await client.post(
                                f"{settings.auth_service_url}/admin/users",
                                headers={"Authorization": f"Bearer {token}"},
                                json={
                                    "email": email,
                                    "password": password,
                                    "full_name": full_name,
                                    "first_name": data.responsible_first_name,
                                    "last_name": data.responsible_last_name,
                                    "role": None,
                                    "role_ids": [],
                                    "is_active": True,
                                },
                            )
                            if resp.status_code != status.HTTP_201_CREATED:
                                raise HTTPException(
                                    status_code=status.HTTP_502_BAD_GATEWAY,
                                    detail="auth_service_error",
                                )
                            created = resp.json()
                            if not isinstance(created, dict) or not isinstance(created.get("id"), int):
                                raise HTTPException(
                                    status_code=status.HTTP_502_BAD_GATEWAY,
                                    detail="auth_service_invalid_response",
                                )
                            responsible_id = int(created["id"])

                        if responsible_id is not None:
                            user_email_to_id[email] = responsible_id

                    barcode_id = data.barcode_id
                    normalized_value = _normalize_barcode(data.barcode_data_12)

                    if barcode_id is not None:
                        existing_by_id = db.get(Barcode, barcode_id)
                        if existing_by_id is None:
                            if normalized_value:
                                barcode_id = None
                            else:
                                raise HTTPException(
                                    status_code=status.HTTP_404_NOT_FOUND,
                                    detail="barcode_id_not_found",
                                )

                    if barcode_id is None and normalized_value:
                        existing = db.execute(
                            select(Barcode).where(Barcode.value == normalized_value)
                        ).scalar_one_or_none()
                        if existing is not None:
                            barcode_id = existing.id
                        else:
                            created = barcode_service.create_barcode(
                                BarcodeCreate(
                                    value=normalized_value[:12]
                                    if len(normalized_value) == 13
                                    else normalized_value,
                                    title=data.title,
                                ),
                                db,
                                commit=False,
                            )
                            barcode_id = created.id

                    if barcode_id is not None:
                        conflict = (
                            db.execute(
                                select(InventoryItem.id).where(
                                    InventoryItem.barcode_id == barcode_id
                                )
                            ).scalar_one_or_none()
                        )
                        if conflict is not None:
                            raise HTTPException(
                                status_code=status.HTTP_409_CONFLICT,
                                detail="barcode_already_in_use",
                            )

                    payload = InventoryItemCreate(
                        title=data.title or "",
                        description=data.description,
                        category=data.category,
                        location_id=location_id,
                        responsible_id=responsible_id,
                        status=data.status,
                        barcode_id=barcode_id,
                    )

                    item = InventoryItem(**payload.model_dump())
                    if data.id is not None:
                        item.id = data.id
                    db.add(item)
                    db.flush()

                    if barcode_id is not None:
                        item.barcode_id = barcode_id
                        db.add(
                            InventoryItemBarcode(
                                inventory_item_id=item.id, barcode_id=barcode_id
                            )
                        )
                    else:
                        generated = barcode_service.create_barcode(
                            BarcodeCreate(value=None, title=item.title),
                            db,
                            commit=False,
                        )
                        item.barcode_id = generated.id
                        db.add(
                            InventoryItemBarcode(
                                inventory_item_id=item.id, barcode_id=generated.id
                            )
                        )

                    db.commit()
                    db.refresh(item)
                    created_item_ids.append(item.id)
                    yield _sse(
                        "row",
                        {
                            "index": index,
                            "total": total,
                            "row_number": row.row_number,
                            "result": "created",
                            "item_id": item.id,
                            "title": item.title,
                        },
                    )
                except HTTPException as exc:
                    db.rollback()
                    error_rows.append(
                        {
                            "row_number": row.row_number,
                            "detail": exc.detail,
                            "status_code": exc.status_code,
                        }
                    )
                    yield _sse(
                        "row",
                        {
                            "index": index,
                            "total": total,
                            "row_number": row.row_number,
                            "result": "error",
                            "detail": exc.detail,
                            "status_code": exc.status_code,
                            "title": data.title,
                        },
                    )
                except Exception:
                    db.rollback()
                    error_rows.append(
                        {"row_number": row.row_number, "detail": "unexpected_error"}
                    )
                    yield _sse(
                        "row",
                        {
                            "index": index,
                            "total": total,
                            "row_number": row.row_number,
                            "result": "error",
                            "detail": "unexpected_error",
                            "title": data.title,
                        },
                    )

        result = InventoryImportConfirmResponse(
            created_count=len(created_item_ids),
            skipped_count=skipped_count,
            error_count=len(error_rows),
            created_item_ids=created_item_ids,
            errors=error_rows,
        )
        yield _sse("done", result.model_dump())

    return _gen()
